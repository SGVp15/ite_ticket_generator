import os
import random

import openpyxl
from openpyxl import load_workbook

from Questions.question import Question
from config import FILE_XLSX, MAP_EXCEL


def read_excel_file(filename='', sheet_names=()) -> {list}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    all_data = {}
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        all_data[sheet_name] = data
    workbook.close()
    return all_data


def get_all_questions_from_excel_file(exam: str) -> [Question]:
    file = FILE_XLSX[exam]

    # data = read_excel_file(file, sheet_names=('SCM',))
    # labels = data['SCM'][0]

    wb = openpyxl.load_workbook(filename=f'{file}', data_only=True)
    page_name = wb.sheetnames
    page_name = str(page_name[0])

    col_id_question = MAP_EXCEL['Код вопроса']
    col_box_question = MAP_EXCEL['Блок вопросов']
    column_enable_question = MAP_EXCEL['Действующий 1-да, 0-нет']
    col_category = MAP_EXCEL['Раздел курса']

    column_a = 'a'
    column_main = 'b'

    all_questions = []
    num_q = 0

    i = 0
    while i < 20_000:
        i += 1
        a = _read_excel(wb, page_name, column_a, i + 1).lower()
        b = _read_excel(wb, page_name, column_a, i + 2).lower()
        c = _read_excel(wb, page_name, column_a, i + 3).lower()
        d = _read_excel(wb, page_name, column_a, i + 4).lower()

        if a in ('a', 'а') and b in ('b', 'в') and c in ('c', 'с') and d == 'd':
            enable_question = _read_excel(wb, page_name, column_enable_question, i)
            if enable_question:
                id_question = _read_excel(wb, page_name, col_id_question, i)
                box_question = _read_excel(wb, page_name, col_box_question, i)
                category = _read_excel(wb, page_name, col_category, i)
                if box_question in ('None', None, 0, 'Nan'):
                    box_question = random.randint(200, 300_000_000_000)
                ans_a = _read_excel(wb, page_name, column_main, i + 1)
                ans_b = _read_excel(wb, page_name, column_main, i + 2)
                ans_c = _read_excel(wb, page_name, column_main, i + 3)
                ans_d = _read_excel(wb, page_name, column_main, i + 4)
                num_q += 1
                text_question = _read_excel(wb, page_name, column_main, i)
                num_question, num_category, version = get_num_question_category_version(id_question)
                all_questions.append(
                    Question(
                        id_question=id_question,
                        text_question=text_question,
                        ans_a=ans_a,
                        ans_b=ans_b,
                        ans_c=ans_c,
                        ans_d=ans_d,
                        box_question=box_question,
                        num_question=num_question,
                        num_category=num_category,
                        version=version,
                        category=category,
                        exam=exam,
                    ))
            i += 3

    # file_json = os.path.join(PATH_QUESTIONS, f'{exam}.json')
    # with open(file_json, 'w', encoding='utf-8') as f:
    #     f.write(json.dumps(all_questions))
    return all_questions


def _read_excel(excel, page_name, column, row):
    sheet_ranges = excel[page_name]
    return str(sheet_ranges[f'{column}{row}'].value).strip()


def get_num_question_category_version(s) -> (int, int, int):
    s = s.split('.')
    try:
        num_question = int(s[1])
        category = int(s[2])
        version = int(s[3])
        return num_question, category, version
    except (IndexError, ValueError):
        return 0, 0, 0


class excel():
    def read_excel(excel, page_name, column, row):
        sheet_ranges = excel[page_name]
        return str(sheet_ranges[f'{column}{row}'].value)


def all_in_one_excel():
    excel = [x for x in os.listdir(f'./') if x.endswith('to_excel.txt')]
    # print(excel)

    out = ['' for _ in range(100)]
    for file in excel:
        with open(file, mode='r', encoding='utf-8') as f:
            s = f.read().split('\n')
            for i in range(len(s)):
                out[i] = out[i] + s[i] + '\t'
    with open(f'./all_excel.txt', mode='w', encoding='utf-8') as f:
        f.write('')
    with open(f'./all_excel.txt', mode='w', encoding='utf-8') as f:
        s = '\n'.join(out)
        f.write(s.strip())
