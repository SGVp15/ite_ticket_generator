import os

import openpyxl

from Questions.Question import Question
from Word.doc_ticket import get_random_index_list_quest


def create_excel_file_for_ispring(questions: [Question]):
    categories = []
    for q in questions:
        if q.category not in categories:
            categories.append(q.category)
    os.makedirs(f'./Category/{questions[0].exam}', exist_ok=True)
    category_file = f'./Category/{questions[0].exam}/category.txt'
    with open(category_file, 'w') as f:
        f.write('')

    with open(category_file, 'a', encoding='utf-8') as f:
        for category in categories:
            f.write(category + '\n')

    questions_by_category = {}
    for category in categories:
        box_questions = {}
        for q in questions:
            if q.category == category:
                if q.box_question not in box_questions.keys():
                    box_questions[q.box_question] = [q]
                else:
                    box_questions[q.box_question].append(q)
        questions_by_category[category] = box_questions.copy()

    for category, list_quest in questions_by_category.items():
        print(f'{category}\t\t{len(list_quest)}')

    head = read_template_ispring()

    max_group = 0
    for category, list_quest in questions_by_category.items():
        for k, v in list_quest.items():
            max_group = max(max_group, len(v))

    for category, list_quest in questions_by_category.items():
        list_len_group = [len(i) for i in list_quest.values()]
        index_list_quest = get_random_index_list_quest(list_len_group, max_group)
        for group_number in range(max_group):
            l_quest = []
            l_questions = [v for v in list_quest.values()]
            for i, v in enumerate(index_list_quest[group_number]):
                l_quest.append(l_questions[i][v])

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            j = 0
            for i, v in enumerate(head):
                worksheet.cell(row=1, column=i + 1, value=str(v))

            # Create a list of values to write to the Excel file
            for q in l_quest:
                j += 1
                values_to_write = ['MC', q.text_question, '', '', '',
                                   f'*{q.ans_a}', q.ans_b, q.ans_c, q.ans_d,
                                   '', '', '', '', '', '', '', '', 1]

                # Write the numbers to the worksheet
                for i, v in enumerate(values_to_write):
                    worksheet.cell(row=j + 1, column=i + 1, value=str(v))

            # Save the workbook to a file
            os.makedirs(name=f'./Category/{questions[0].exam}/{group_number}', exist_ok=True)
            workbook.save(f'./Category/{questions[0].exam}/{group_number}/{category[:2]}.xlsx')


def read_template_ispring(file='./import_ispring.xlsx'):
    if os.path.isfile(file):
        workbook = openpyxl.load_workbook(file)
        page_name = workbook.sheetnames
        worksheet = workbook[page_name[0]]
        for row in worksheet.iter_rows(values_only=True):
            return row
